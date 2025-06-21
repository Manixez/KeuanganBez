from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Updater, CommandHandler, CallbackQueryHandler, MessageHandler, Filters, CallbackContext
from openpyxl import load_workbook
from datetime import datetime
import pandas as pd
import pytz
import os
from apscheduler.schedulers.background import BackgroundScheduler

# Ganti dengan token dari @BotFather
TOKEN = os.getenv("BOT_TOKEN")

# Path file Excel (pastikan file-nya ada di folder ini)
EXCEL_PATH = 'Template_Keuangan_TelegramBot.xlsx'

# Ganti ini dengan chat_id kamu (lihat dari /start response)
YOUR_CHAT_ID = int(os.getenv("CHAT_ID"))  # <-- GANTI dengan chat_id kamu

# Menyimpan status user sementara
user_state = {}

# ========= HANDLER: /start =========
def start(update: Update, context: CallbackContext):
    keyboard = [
        [InlineKeyboardButton("✅ Pemasukan", callback_data='pemasukan'),
         InlineKeyboardButton("❌ Pengeluaran", callback_data='pengeluaran')],
        [InlineKeyboardButton("📆 Mingguan", callback_data='mingguan'),
         InlineKeyboardButton("🗓️ Bulanan", callback_data='bulanan')]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    update.message.reply_text("Halo! Pilih menu:", reply_markup=reply_markup)

# ========= HANDLER: Tombol Inline =========
def button_handler(update: Update, context: CallbackContext):
    query = update.callback_query
    query.answer()
    chat_id = query.message.chat_id

    if query.data == 'pemasukan':
        user_state[chat_id] = 'awaiting_income'
        query.message.reply_text("Masukkan nominal pemasukan kamu:")

    elif query.data == 'pengeluaran':
        user_state[chat_id] = {'state': 'choosing_expense_category'}
        keyboard = [
            [InlineKeyboardButton("🍚 Bahan Pokok", callback_data='kat_bahan')],
            [InlineKeyboardButton("🍱 Makanan Pokok", callback_data='kat_makanan')],
            [InlineKeyboardButton("🍩 Jajanan", callback_data='kat_jajanan')],
            [InlineKeyboardButton("🏫 Kampus", callback_data='kat_kampus')],
            [InlineKeyboardButton("🎓 Himpunan", callback_data='kat_himpunan')],
            [InlineKeyboardButton("⛽ Bensin", callback_data='kat_bensin')]
        ]
        query.message.reply_text("Pilih kategori pengeluaran:", reply_markup=InlineKeyboardMarkup(keyboard))

    elif query.data.startswith('kat_'):
        kategori = query.data.replace('kat_', '').capitalize()
        kategori = kategori.replace("bahan", "Bahan Pokok").replace("makanan", "Makanan Pokok")
        kategori = kategori.replace("jajanan", "Jajanan").replace("kampus", "Pengeluaran Kampus")
        kategori = kategori.replace("himpunan", "Pengeluaran Himpunan").replace("bensin", "Bensin")

        user_state[chat_id] = {'state': 'awaiting_expense', 'kategori': kategori}
        query.message.reply_text(f"Masukkan nominal pengeluaran untuk kategori *{kategori}*:", parse_mode="Markdown")

    elif query.data == 'mingguan':
        send_weekly_report(chat_id, context)

    elif query.data == 'bulanan':
        send_monthly_report(chat_id, context)

# ========= HANDLER: Pesan (Nominal) =========
def message_handler(update: Update, context: CallbackContext):
    chat_id = update.message.chat_id
    text = update.message.text

    if chat_id in user_state:
        state_data = user_state[chat_id]
        state = state_data if isinstance(state_data, str) else state_data.get('state')

        if state == 'awaiting_income':
            try:
                nominal = int(text)
                today = datetime.today().date()
                minggu_ke = today.isocalendar().week
                bulan = today.strftime('%B')

                wb = load_workbook(EXCEL_PATH)
                ws = wb['Transaksi']
                ws.append([today, nominal, None, None, minggu_ke, bulan])
                wb.save(EXCEL_PATH)

                update.message.reply_text(f"Pemasukan sebesar Rp{nominal:,} berhasil dicatat!")
                user_state.pop(chat_id)

            except ValueError:
                update.message.reply_text("Input harus berupa angka ya. Coba lagi!")

        elif state == 'awaiting_expense':
            try:
                nominal = int(text)
                kategori = state_data.get('kategori')
                today = datetime.today().date()
                minggu_ke = today.isocalendar().week
                bulan = today.strftime('%B')

                wb = load_workbook(EXCEL_PATH)
                ws = wb['Transaksi']
                ws.append([today, None, nominal, kategori, minggu_ke, bulan])
                wb.save(EXCEL_PATH)

                update.message.reply_text(f"Pengeluaran Rp{nominal:,} untuk *{kategori}* berhasil dicatat!", parse_mode="Markdown")
                user_state.pop(chat_id)

            except ValueError:
                update.message.reply_text("Input harus berupa angka ya. Coba lagi!")

# ========= LAPORAN MINGGUAN =========
def send_weekly_report(chat_id, context: CallbackContext):
    try:
        df = pd.read_excel(EXCEL_PATH, sheet_name='Transaksi')
        today = datetime.today().date()
        minggu_ini = today.isocalendar().week

        df_minggu = df[df['Minggu Ke'] == minggu_ini]

        pemasukan = df_minggu['Pemasukan'].sum(skipna=True)
        pengeluaran = df_minggu['Pengeluaran'].sum(skipna=True)
        saldo = pemasukan - pengeluaran

        response = (
            f"📊 *Laporan Mingguan - Minggu ke-{minggu_ini}*\n\n"
            f"💰 Total Pemasukan: Rp{pemasukan:,.0f}\n"
            f"💸 Total Pengeluaran: Rp{pengeluaran:,.0f}\n"
            f"🧾 Saldo: Rp{saldo:,.0f}"
        )
        context.bot.send_message(chat_id=chat_id, text=response, parse_mode='Markdown')

    except Exception as e:
        context.bot.send_message(chat_id=chat_id, text=f"Gagal membuat laporan mingguan.\nError: {e}")

# ========= LAPORAN BULANAN =========
def send_monthly_report(chat_id, context: CallbackContext):
    try:
        df = pd.read_excel(EXCEL_PATH, sheet_name='Transaksi')
        bulan_ini = datetime.today().strftime('%B')

        df_bulan = df[df['Bulan'] == bulan_ini]

        pemasukan = df_bulan['Pemasukan'].sum(skipna=True)
        pengeluaran = df_bulan['Pengeluaran'].sum(skipna=True)
        saldo = pemasukan - pengeluaran

        response = (
            f"📅 *Laporan Bulanan - {bulan_ini}*\n"
            f"💰 Total Pemasukan: Rp{pemasukan:,.0f}\n"
            f"💸 Total Pengeluaran: Rp{pengeluaran:,.0f}\n"
            f"🧾 Saldo: Rp{saldo:,.0f}"
        )
        context.bot.send_message(chat_id=chat_id, text=response, parse_mode='Markdown')

    except Exception as e:
        context.bot.send_message(chat_id=chat_id, text=f"Gagal membuat laporan bulanan.\nError: {e}")

# ========= SCHEDULER =========
def schedule_weekly_report(context: CallbackContext):
    send_weekly_report(YOUR_CHAT_ID, context)

def schedule_monthly_report(context: CallbackContext):
    send_monthly_report(YOUR_CHAT_ID, context)

# ========= MAIN =========
def main():
    updater = Updater(TOKEN)
    dp = updater.dispatcher

    dp.add_handler(CommandHandler("start", start))
    dp.add_handler(CallbackQueryHandler(button_handler))
    dp.add_handler(MessageHandler(Filters.text & ~Filters.command, message_handler))

    scheduler = BackgroundScheduler(timezone=pytz.timezone("Asia/Jakarta"))
    scheduler.add_job(schedule_weekly_report, trigger='cron', day_of_week='sun', hour=20, minute=0, args=[updater.bot])
    scheduler.add_job(schedule_monthly_report, trigger='cron', day=1, hour=7, minute=0, args=[updater.bot])
    scheduler.start()

    updater.start_polling()
    print("Bot berjalan... tekan Ctrl+C untuk berhenti.")
    updater.idle()

if __name__ == '__main__':
    main()
